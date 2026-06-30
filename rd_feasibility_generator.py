#!/usr/bin/env python3
"""
rd_feasibility_generator.py
Generates an R&D Tax Credit Feasibility Study HTML report.

Reads:
  - Output/rd_credit_calculator_v2_scenario_*.xlsx  — scenario results
  - Input/logo.png  — logo

Writes:
  - Output/rd_feasibility_<Company>_<Year>.html

Usage:
    python rd_feasibility_generator.py
"""

import sys, os, re, base64, glob
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl

try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False


# ---------------------------------------------------------------------------
# Cover letter template  (edit freely)
# ---------------------------------------------------------------------------

LETTER_PARAS = [
    "[study_date]",
    "EIN: [EIN_number]",
    "[Company_Name] Research & Development Tax Credit Feasibility Study",
    "Enclosed is the preliminary R&D tax credit feasibility study prepared for your review. This report "
    "provides a high-level overview of the Research and Development (R&D) Tax Credit under Section 41, a summary "
    "of qualifying activities and expenses, and preliminary credit estimates based on the information "
    "provided. This study is intended to assist in planning and to support a discussion of the potential "
    "federal R&D tax credit benefit. The estimated credit benefit will depend on the company’s actual qualified "
    "research expenses, gross receipts history, prior-year R&D expenses, available documentation, "
    "entity type, tax liability, and other applicable limitations.",
    "",
    "Table of Contents:",
    "",
    "1. R&D Credit Overview",
    "2. Qualifying and Non-Qualifying Activities",
    "3. Qualified Research Expenses Review",
    "4. Supporting Documentation for Qualified Research Expenses",
    "5. IRS Scrutiny and Management Responsibility",
    "6. Potential Tax Savings",
    "7. Important Limitations and Disclaimers",
    "8. Scenario Analysis",
]

DOLLAR_FIELDS = {"qre_wages", "qre_supplies", "qre_contract_us", "rd_credit_amount"}

# ---------------------------------------------------------------------------
# Content pages — edit section text directly here
#
# Structure: list of pages. Each page is a dict:
#   {
#     "page_title": str or None,   # optional gray divider shown above cards
#     "compact":    bool,          # True = tighter spacing (use for 7.1–7.4 page)
#     "sections":   [              # list of (heading, [bullet, ...]) tuples
#       ("Heading Text", ["bullet 1", "bullet 2", ...]),
#       ...
#     ]
#   }
# ---------------------------------------------------------------------------

CONTENT_PAGES = [
    # ── Page 1 ────────────────────────────────────────────────────────────
    {
        "page_title": "1. R&D Credit Overview",
        "compact": False,
        "sections": [
            (
                "1.1 R&D Credit Overview",
                [
                    "The federal R&D tax credit is available to companies that invest time and resources "
                    "into developing or improving products, software, processes, techniques, formulas, or "
                    "other business components.",

                    "Technology, software, manufacturing, engineering, product-development, and "
                    "process-improvement companies may qualify when the work involves technical uncertainty "
                    "and experimentation.",

                    "The activity does not have to result in a commercially successful product. Abandoned "
                    "projects, prototypes, redesigns, and iterative testing may still be relevant if the "
                    "work was intended to resolve technical uncertainty.",

                    "The credit can reduce federal income tax dollar-for-dollar and, for certain qualified "
                    "small businesses, may also be available as a payroll tax credit election.",
                ],
            ),
            (
                "1.2 Payroll Tax Credit Option",
                [
                    "If the company has limited taxable income or is operating at a loss, the R&D credit "
                    "may still provide a benefit if the company qualifies for the payroll tax credit election.",

                    "The maximum payroll tax credit election is generally $500,000 per year, subject to "
                    "applicable limitations.",

                    "This option is only available if the company qualifies as a Qualified Small Business (QSB) for the "
                    "payroll tax election and must be coordinated with the income tax return and payroll tax filings.",

                ],
            ),
            (
                "1.3 Domestic R&D Deduction and Section 280C Considerations",
                [
                    "For tax years beginning after December 31, 2024, domestic research or experimental "
                    "costs may be currently deducted under Section 174A, subject to applicable rules and elections.",
                    "The R&D credit calculation is separate from the deduction calculation, but the two "
                    "items interact for tax purposes.",
                    "Using the Section 280C reduced credit approach can simplify the presentation by "
                    "reducing the credit amount instead of requiring a separate reduction to the related deduction.",
                    "In general, the advantage after the recent law change is that qualifying domestic R&D "
                    "costs may provide both a current deduction and a reduced R&D credit, subject to the "
                    "company's facts and limitations.",
                    "State tax treatment may be different from federal treatment and should be reviewed separately.",
                ],
            ),
        ],
    },

    # ── Page 2 ────────────────────────────────────────────────────────────
    {
        "page_title": "2. Qualifying and Non-Qualifying Activities",
        "compact": False,
        "sections": [
            (
                "2.1 Activities That Qualify",
                [
                    "Developing new or improved software, platforms, applications, tools, or internal systems.",
                    "Improving product functionality, performance, reliability, quality, or efficiency.",
                    "Testing different technical approaches, designs, formulas, prototypes, or engineering methods.",
                    "Resolving technical uncertainty through trial and error, modeling, testing, debugging, "
                    "or iterative development.",
                    "Designing or improving manufacturing, production, or engineering processes.",
                ],
            ),
            (
                "2.2 Activities That Generally Do Not Qualify",
                [
                    "Routine maintenance or cosmetic changes that do not involve technical uncertainty.",
                    "Administrative, sales, marketing, customer support, or general management activities.",
                    "Training, implementation, or data entry unrelated to technical development.",
                    "Research conducted after the product or process is ready for commercial use, unless "
                    "additional qualified development is performed.",
                ],
            ),
        ],
    },

    # ── Page 3 ────────────────────────────────────────────────────────────
    {
        "page_title": "3. Qualified Research Expenses Review",
        "compact": False,
        "sections": [
            (
                "3.1 Qualified Research Expenses:",
                [
                    "Wages - paid to employees directly supervising or directly supporting qualified research.",
                    "Supplies  - used or consumed in the research process.",
                    "U.S. contractors - eligible payments to  performing qualified research on "
                    "behalf of the company.Commonly, only 65% of eligible U.S. contractor costs are included in"
                    "qualified research expenses.",
                    "Computer rental -  cloud computing, development environment, testing environment, "
                    "model-training, or usage-based AI compute costs, general hosting"

                ],
            ),
            (
                "3.2 Foreign R&D Contractors",
                [
                    "Foreign contractor costs are generally not eligible for the federal Section 41 R&D credit.",
                    "Foreign R&D costs generally must be capitalized and amortized over 15 years, which can "
                    "significantly reduce the current-year deduction.",
                    "For depreciation calculations, we need contractor location, amount paid, estimated time spent on "
                    "qualified research, and a description of the work performed.",
                ],
            ),
        ],
    },

    # ── Page 4 — Documentation 7.1–7.4 (compact to fit on one page) ──────
    {
        "page_title": "4. Supporting Documentation for Qualified Research Expenses",
        "compact": True,
        "sections": [
            (
                "4.1 General Project Support and Company Description",
                [
                    "Brief description of the company, its products or services, and the business "
                    "components developed or improved during the year.",
                    "Project descriptions explaining what was developed or improved.",
                    "Description of the technical uncertainty or challenge the company was trying to resolve.",
                    "Evidence of alternatives evaluated, testing performed, and changes made during development.",
                    "Technical notes, design records, engineering reports, product roadmaps, tickets, "
                    "version history, GitHub/Jira records, prototypes, test results, or project timelines.",
                ],
            ),
            (
                "4.2 Documentation: Employee Wages",
                [
                    "Employee names, job titles, departments, and role descriptions.",
                    "Description of each employee’s involvement in qualified research projects.."
                    "Payroll records supporting total wages paid during the year, including W-2 wage "
                    "information when applicable.",
                    "Time tracking, project tracking, management estimates, or other records supporting "
                    "the percentage of time spent on qualified activities.",
                    "Management should be able to explain and support how wage allocations were determined.",
                ],
            ),
            (
                "4.3 Documentation: Supplies",
                [
                    "Invoices, receipts, purchase records, or general ledger detail for supplies included "
                    "in the claim.",
                    "Description of how the supplies were used or consumed in qualified research activities.",
                    "Exclude general office supplies, capital assets, equipment, or costs not directly "
                    "tied to qualified research.",
                ],
            ),
            (
                "4.4 Documentation: U.S. Contractors",
                [
                    "Signed contracts, statements of work, proposals, invoices, and payment records.",
                    "Description of the research or development work performed by the contractor.",
                    "Support showing the contractor performed technical development, testing, design, or "
                    "experimentation rather than routine services.",
                    "Confirmation that the work was performed in the United States when claimed as U.S. "
                    "contract research.",
                ],
            ),
        ],
    },

    # ── Page 5 ────────────────────────────────────────────────────────────
    {
        "page_title": "5. IRS Scrutiny and Management Responsibility",
        "compact": False,
        "sections": [
            (
                "5. IRS Scrutiny and Management Responsibility",
                [
                    "R&D credit claims have received increased IRS attention, especially where claims are "
                    "based on broad estimates, unsupported payroll percentages, or limited project documentation.",
                    "CFO Associates will review the information provided, ask follow-up questions, perform "
                    "reasonableness checks, and prepare the calculation based on available facts.",
                    "The company must maintain records in sufficient detail to support the activities, "
                    "projects, employees, contractors, and expenses included in the claim.",
                    "Management is ultimately responsible for confirming that the activities qualify and "
                    "that adequate documentation exists to support the credit if the IRS examines the return.",
                    "The strongest R&D claims are supported by project-level records, employee involvement, "
                    "contractor support, and a clear connection between the technical work and the costs claimed.",
                ],
            ),
        ],
    },

    # ── Page 6 ────────────────────────────────────────────────────────────
    {
        "page_title": "6. Potential Tax Savings",
        "compact": False,
        "sections": [
            (
                "6.1 C Corporation Tax Impact",
                [
                    "For a C corporation, the federal R&D credit generally reduces corporate income tax "
                    "dollar-for-dollar, subject to applicable limitations.",
                    "The regular federal corporate income tax rate is 21%, so deductible expenses generally "
                    "reduce tax at approximately 21 cents per dollar before considering other limitations.",
                    "Foreign R&D amortization can reduce the current-year deduction and may increase taxable "
                    "income compared with book expenses.",
                    "The net benefit should be modeled by comparing the credit benefit, deduction impact, "
                    "taxable income, net operating losses, and any payroll tax credit election.",
                ],
            ),
            (
                "6.2 S Corporation Tax Impact",
                [
                    "For an S corporation, the R&D credit generally flows through to the shareholders and "
                    "is reported on their individual income tax returns.",
                    "The company-level calculation determines the credit amount, but the actual tax savings "
                    "is determined at the shareholder level.",
                    "The benefit depends on the shareholders' personal tax situation, income tax liability, "
                    "basis, passive activity limitations, AMT considerations, and other credit limitations.",
                    "If the company qualifies for the payroll tax credit election, part of the credit may "
                    "be used against payroll taxes instead of, or before, shareholder-level income tax benefit.",
                ],
            ),
        ],
    },

    # ── Page 7 ────────────────────────────────────────────────────────────
    {
        "page_title": "7. Important Limitations and Disclaimers",
        "compact": False,
        "sections": [
            (
                "7.1 Important Limitations and Disclaimers",
                [
                    "All numbers shown in this presentation are estimates unless otherwise stated.",
                    "Actual tax savings may differ based on final qualified research expenses, gross "
                    "receipts, prior-year data, tax liability, entity type, ownership, payroll election "
                    "eligibility, net operating losses, and Section 174/174A treatment.",
                    "A credit estimate does not guarantee that the company can use the full credit in "
                    "the current year.",
                    "The R&D credit should be claimed only when the company has a reasonable basis and "
                    "adequate records to support both the activities and the related costs.",
                    "This presentation is for planning and discussion purposes and should be finalized "
                    "after review of the company's actual facts and records.",
                ],
            ),
        ],
    },
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _num(v):
    if v is None or v == "": return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    try: return float(s)
    except ValueError: return 0.0

def esc(text):
    return (str(text)
            .replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))

def apply_wildcards(text, fields):
    def replacer(m):
        key = m.group(1).lower()
        val = fields.get(key)
        if val is None: return m.group(0)
        if isinstance(val, datetime): return val.strftime("%B %d, %Y")
        if key in DOLLAR_FIELDS:
            try: return f"${float(val):,.2f}"
            except: return str(val)
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(val)
    return re.sub(r'\[([A-Za-z_]+)\]', replacer, text)


# ---------------------------------------------------------------------------
# Data readers
# ---------------------------------------------------------------------------

def load_company_info(scenario_paths: list) -> dict:
    """Pull company metadata from first scenario's Input Data sheet."""
    if not scenario_paths:
        return {}
    wb = openpyxl.load_workbook(scenario_paths[0], data_only=True)
    if "Input Data" not in wb.sheetnames:
        return {}
    ws = wb["Input Data"]
    key_map = {
        "Company Name":          "company_name",
        "Entity Type":           "entity_type",
        "Tax Year":              "tax_year",
        "EIN":                   "ein_number",
        "Study Date":            "study_date",
        "Preparer Name":         "preparer_name",
        "Year Company Started":  "year_started",
        "Year Research Started": "year_research_started",
    }
    info = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        field, value = row[0], row[1]
        if field and field in key_map:
            info[key_map[field]] = value
    info["EIN_number"]   = info.get("ein_number", "")
    info["Company_Name"] = info.get("company_name", "")
    return info


def load_scenario(path: str) -> dict:
    """Read Calculated Results sheet from a scenario xlsx."""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    if "Calculated Results" not in wb.sheetnames:
        return result
    ws = wb["Calculated Results"]
    warnings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        label, value = row[0], row[1]
        if label:
            result[label] = value
        elif value and label is None:
            warnings.append(str(value))
    result["_warnings"] = warnings
    return result


# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------

CSS = """
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

body {
    font-family: Verdana, Geneva, Tahoma, sans-serif;
    font-size: 9pt;
    color: #1a1a1a;
    background: #d8dde6;
    padding: 1.2rem 0 2rem 0;
}

.page {
    width: 8.5in;
    min-height: 11in;
    margin: 0 auto 1.2rem auto;
    background: #fff;
    box-shadow: 0 4px 20px rgba(0,0,0,.22);
    display: flex;
    flex-direction: column;
}

@media print {
    body { background: #fff; padding: 0; }
    * { -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
    .page {
        box-shadow: none; margin: 0; width: 100%;
        page-break-after: always; page-break-inside: avoid;
    }
    .page:last-child { page-break-after: avoid; }
}
@page { size: 8.5in 11in; margin: 0; }

/* ── Page chrome ── */
.page-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0.32rem 0.9in;
    background: #003366;
    color: #fff;
    font-size: 7.5pt;
    font-weight: bold;
    letter-spacing: 0.3px;
    flex-shrink: 0;
}
.page-header span { flex: 1; }
.page-header span:nth-child(2) { text-align: center; }
.page-header span:last-child   { text-align: right; }

.page-content {
    flex: 1;
    padding: 0.5in 0.9in 0.3in 0.9in;
    overflow: hidden;
}

.page-footer {
    flex-shrink: 0;
    padding: 0 0.9in 0.28in 0.9in;
}
.footer-rule { border: none; border-top: 1px solid #b8c8d8; margin-bottom: 0.2rem; }
.footer-text  { font-size: 7pt; color: #888; text-align: center; letter-spacing: 0.5px; }

/* ── Cover page ── */
.cover-content {
    flex: 1;
    padding: 0.6in 0.9in 0.4in 0.9in;
    display: flex;
    flex-direction: column;
}
.cover-logo { height: 264px; margin-bottom: 0.45in; }
.cover-title-block {
    text-align: center;
    border-bottom: 3px solid #003366;
    padding-bottom: 1.2rem;
    margin-bottom: 1.2rem;
}
.cover-title-block h1 {
    font-size: 17pt; font-weight: bold;
    color: #003366; letter-spacing: 0.5px; line-height: 1.25;
}
.cover-title-block h2 { font-size: 26pt; font-weight: bold; color: #003366; margin-top: 0.4rem; }
.cover-title-block h3 { font-size: 13pt; color: #003366; margin-top: 0.2rem; font-weight: bold; }
.cover-bottom-info {
    margin-top: auto;
    text-align: right;
    font-size: 9pt;
    color: #555;
    line-height: 1.8;
}

/* ── Cover letter ── */
.cover-letter { flex: 1; }
.letter-date  { font-size: 9pt; margin: 0.4rem 0; }
.letter-re    { font-size: 9pt; font-weight: bold; margin: 0.5rem 0 0.7rem; }
.cover-letter p {
    font-size: 9pt; line-height: 1.65;
    margin-bottom: 0.6rem; text-align: justify; color: #1a1a1a;
}
.toc-sub {
    padding-left: 1.2rem;
    font-weight: bold;
    margin-bottom: 0.25rem !important;
    color: #003366;
}

/* ── Page section divider (e.g. "7. Documentation") ── */
.page-section-divider {
    font-size: 11pt;
    font-weight: bold;
    color: #003366;
    border-bottom: 2px solid #d0daea;
    padding-bottom: 0.25rem;
    margin-bottom: 0.6rem;
}

/* ── Info cards with teal bullets ── */
.card-header {
    background: #003366;
    color: #fff;
    font-size: 9.5pt;
    font-weight: bold;
    padding: 0.3rem 0.75rem;
    border-radius: 5px 5px 0 0;
    letter-spacing: 0.2px;
    margin-top: 1.1rem;    /* spacing between sections */
}
.card-header:first-child,
.page-section-divider + .card-header {
    margin-top: 0;         /* no top gap for first card on page */
}

.bullet-list {
    background: #f5f8fc;
    border: 1px solid #d0daea;
    border-top: none;
    border-radius: 0 0 5px 5px;
    padding: 0.45rem 0.75rem 0.4rem 0.75rem;
    margin-bottom: 0;
}

.bullet-item {
    display: flex;
    align-items: flex-start;
    margin-bottom: 0.42rem;
    font-size: 9.5pt;
    line-height: 1.5;
    gap: 0.5rem;
}
.bullet-item:last-child { margin-bottom: 0.1rem; }

.bullet-dot {
    display: inline-block;
    width: 11px;
    height: 11px;
    min-width: 11px;
    background: #0e7490;
    border-radius: 50%;
    margin-top: 3px;
}

/* ── Compact group: tighter inter-card spacing for 4.1–4.4 page ── */
.compact-group .card-header {
    margin-top: 0.45rem;
}
.compact-group .card-header:first-child { margin-top: 0; }
.compact-group .bullet-item {
    margin-bottom: 0.28rem;
    line-height: 1.45;
}
.compact-group .bullet-list {
    padding: 0.3rem 0.75rem 0.25rem 0.75rem;
}

/* ── Scenario pages ── */
.scenario-title {
    font-size: 14pt;
    font-weight: bold;
    color: #003366;
    border-bottom: 2.5px solid #003366;
    padding-bottom: 0.25rem;
    margin-bottom: 0.5rem;
}

.info-pills {
    display: flex;
    flex-wrap: wrap;
    gap: 0.4rem;
    margin-bottom: 0.55rem;
}
.info-pill {
    background: #eef3fa;
    border: 1px solid #c2d2e8;
    border-radius: 20px;
    padding: 0.2rem 0.7rem;
    font-size: 8pt;
    display: flex;
    gap: 0.35rem;
    align-items: center;
}
.pill-label { color: #555; }
.pill-value { font-weight: bold; color: #003366; }

.qre-grid {
    display: flex;
    gap: 0.4rem;
    flex-wrap: wrap;
    background: #f5f8fc;
    border: 1px solid #d0daea;
    border-top: none;
    border-radius: 0 0 5px 5px;
    padding: 0.4rem 0.6rem;
    margin-bottom: 0.2rem;
}
.qre-card {
    flex: 1;
    min-width: 1.3in;
    background: #fff;
    border: 1px solid #d0daea;
    border-radius: 6px;
    padding: 0.4rem 0.5rem;
    text-align: center;
}
.qre-card.qre-total {
    background: #e8eef8;
    border-color: #aabbd4;
    font-weight: bold;
}
.qre-label { font-size: 7.5pt; color: #555; margin-bottom: 0.2rem; }
.qre-value { font-size: 9.5pt; font-weight: bold; color: #003366; }
.method-row {
    background: #f5f8fc;
    border: 1px solid #d0daea;
    border-top: none;
    border-radius: 0 0 5px 5px;
    padding: 0.35rem 0.75rem;
    font-size: 8.5pt;
    margin-bottom: 0.4rem;
}
.method-label { color: #555; }
.method-value { font-weight: bold; color: #003366; margin-left: 0.3rem; }

.credit-banner {
    background: linear-gradient(135deg, #002855 0%, #0047a0 100%);
    border-radius: 10px;
    padding: 0.9rem 1.2rem;
    text-align: center;
    margin: 0.5rem 0 0.4rem 0;
    box-shadow: 0 3px 12px rgba(0,40,100,.25);
}
.credit-banner-label {
    font-size: 10pt;
    color: #a8c4e8;
    font-weight: bold;
    letter-spacing: 0.5px;
    text-transform: uppercase;
    margin-bottom: 0.35rem;
}
.credit-banner-amount {
    font-size: 30pt;
    font-weight: bold;
    color: #ffffff;
    letter-spacing: 1px;
    line-height: 1.1;
    margin-bottom: 0.3rem;
}
.credit-banner-amount.insufficient {
    font-size: 14pt;
    color: #ffc78a;
}
.tax-savings-row {
    margin-top: 0.4rem;
    padding-top: 0.35rem;
    border-top: 1px solid rgba(255,255,255,0.25);
    display: flex;
    justify-content: center;
    align-items: center;
    gap: 0.5rem;
    font-size: 9pt;
    color: #c8ddf5;
}
.savings-amount {
    font-size: 13pt;
    font-weight: bold;
    color: #7de8a0;
}

.warning-note {
    background: #fff8e1;
    border-left: 3px solid #f0a000;
    padding: 0.3rem 0.6rem;
    font-size: 8pt;
    color: #5a3e00;
    border-radius: 0 4px 4px 0;
    margin-top: 0.35rem;
    line-height: 1.5;
}
"""


# ---------------------------------------------------------------------------
# Page structure helpers
# ---------------------------------------------------------------------------

LOGO_B64 = ""   # populated at runtime

def _page_footer():
    return (
        '<div class="page-footer">'
        '<hr class="footer-rule">'
        '<div class="footer-text">CFO Associates &nbsp;|&nbsp; Confidential — For Discussion Purposes Only</div>'
        '</div>'
    )

def _page_header(fields):
    company  = esc(fields.get("company_name", ""))
    tax_year = esc(str(fields.get("tax_year", "")))
    return (
        '<div class="page-header">'
        '<span>' + company + '</span>'
        '<span>R&amp;D Tax Credit Feasibility Study</span>'
        '<span>Tax Year ' + tax_year + '</span>'
        '</div>'
    )


# ---------------------------------------------------------------------------
# Cover page + letter
# ---------------------------------------------------------------------------

def build_cover_pages(fields):
    company  = esc(fields.get("company_name", ""))
    tax_year = esc(str(fields.get("tax_year", "")))
    preparer = esc(str(fields.get("preparer_name", "")))

    page1 = (
        '<div class="page">'
        '<div class="cover-content">'
        '<img src="' + LOGO_B64 + '" class="cover-logo" alt="CFO Associates">'
        '<div class="cover-title-block">'
        '<h1>Research &amp; Development Tax Credit</h1>'
        '<h3>Feasibility Study</h3>'
        '<h2>' + company + '</h2>'
        '</div>'
        '<div class="cover-bottom-info">'
        '<div>Tax Year ' + tax_year + '</div>'
        '<div>Prepared by ' + preparer + '</div>'
        '</div>'
        '</div>'
        + _page_footer() +
        '</div>\n'
    )

    paras = [apply_wildcards(p, fields) for p in LETTER_PARAS]
    letter_html = '<div class="cover-content"><div class="cover-letter">'
    for idx, para in enumerate(paras):
        if para.startswith("Re:"):
            letter_html += '<p class="letter-re">' + esc(para) + '</p>\n'
        elif idx == 0:
            letter_html += '<p class="letter-date">' + esc(para) + '</p>\n'
        elif para.startswith("<b>"):
            letter_html += '<p class="toc-sub">' + para + '</p>\n'
        else:
            letter_html += '<p>' + esc(para) + '</p>\n'
    letter_html += '</div></div>'

    page2 = '<div class="page">\n' + letter_html + _page_footer() + '</div>\n'
    return page1 + page2


# ---------------------------------------------------------------------------
# Content pages
# ---------------------------------------------------------------------------

def _card_section(heading, bullets, compact=False):
    html = '<div class="card-header">' + esc(heading) + '</div>\n'
    if bullets:
        html += '<div class="bullet-list">\n'
        for b in bullets:
            html += (
                '<div class="bullet-item">'
                '<span class="bullet-dot"></span>'
                '<span>' + esc(b) + '</span>'
                '</div>\n'
            )
        html += '</div>\n'
    return html


def build_content_pages(fields: dict) -> str:
    pages_html = ""
    for page in CONTENT_PAGES:
        page_title = page.get("page_title")
        compact    = page.get("compact", False)
        sections   = page.get("sections", [])

        if not sections:
            continue

        body = ""
        if page_title:
            body += '<div class="page-section-divider">' + esc(page_title) + '</div>\n'

        wrapper_open  = '<div class="compact-group">\n' if compact else ''
        wrapper_close = '</div>\n'                        if compact else ''

        body += wrapper_open
        for heading, bullets in sections:
            body += _card_section(heading, bullets, compact)
        body += wrapper_close

        pages_html += (
            '<div class="page">'
            + _page_header(fields)
            + '<div class="page-content">' + body + '</div>'
            + _page_footer()
            + '</div>\n'
        )
    return pages_html


# ---------------------------------------------------------------------------
# Scenario pages
# ---------------------------------------------------------------------------

def build_scenario_page(result: dict, fields: dict) -> str:
    def get(k):
        return result.get(k)

    def d(v):
        if v is None: return "—"
        if isinstance(v, (int, float)): return f"${v:,.2f}"
        return str(v)

    scenario_num = get("Scenario Number") or "—"
    company      = get("Company Name")    or fields.get("company_name", "")
    tax_year     = get("Tax Year")        or fields.get("tax_year", "")
    entity_type  = get("Entity Type")     or ""

    wages    = get("Wages for qualified services")                        or 0
    supplies = get("Cost of supplies")                                    or 0
    comps    = get("Rental or lease costs of computers")                  or 0
    contract = get("Applicable % of contract research expenses (65%)") or 0
    total    = get("Total qualified research expenses")                   or 0

    regular_val = get("Regular Credit Method — 280C applied (15.8%)")
    asc_val     = get("Alternative Simplified Credit (ASC) — 280C applied")
    recommended = get("Recommended R&D Tax Credit")

    if isinstance(recommended, (int, float)):
        credit_disp  = d(recommended)
        credit_class = "credit-banner-amount"
    else:
        credit_disp  = "Insufficient Data"
        credit_class = "credit-banner-amount insufficient"

    if isinstance(regular_val, (int, float)):
        method_str = "Regular Credit Method — 280C reduced credit (15.8%)"
    elif isinstance(asc_val, (int, float)):
        method_str = "Alternative Simplified Credit (ASC) — 280C applied (feasibility estimate)"
    else:
        method_str = "Insufficient data — see notes below"

    tax_savings_html = ""
    if (str(entity_type).strip() == "C Corporation"
            and isinstance(recommended, (int, float)) and recommended > 0):
        savings = recommended * 0.21
        tax_savings_html = (
            '<div class="tax-savings-row">'
            '<span>Potential Tax Savings (C Corporation @ 21%)</span>'
            '<span class="savings-amount">' + f"${savings:,.2f}" + '</span>'
            '</div>'
        )

    warn_html = ""
    if not isinstance(regular_val, (int, float)) and isinstance(asc_val, (int, float)):
        warn_html = (
            '<div class="warning-note">&#9888; Regular method not calculated — '
            'prior-year gross receipts not provided. ASC used as feasibility estimate. '
            'Results should be confirmed once gross receipts data is available.</div>'
        )

    body = (
        '<div class="page-section-divider">8. Scenario Analysis</div>\n'
        f'<div class="scenario-title">Scenario {esc(str(scenario_num))} &mdash; R&amp;D Tax Credit Estimate</div>\n'
        f'<div class="info-pills">'
        f'<div class="info-pill"><span class="pill-label">Company</span><span class="pill-value">{esc(str(company))}</span></div>'
        f'<div class="info-pill"><span class="pill-label">Tax Year</span><span class="pill-value">{esc(str(tax_year))}</span></div>'
        f'<div class="info-pill"><span class="pill-label">Entity</span><span class="pill-value">{esc(str(entity_type))}</span></div>'
        f'<div class="info-pill"><span class="pill-label">280C Election</span><span class="pill-value">Yes</span></div>'
        f'</div>\n'
        f'<div class="card-header" style="margin-top:0">Qualified Research Expenses (QRE)</div>'
        f'<div class="qre-grid">'
        f'<div class="qre-card"><div class="qre-label">Wages</div><div class="qre-value">{d(wages)}</div></div>'
        f'<div class="qre-card"><div class="qre-label">Supplies</div><div class="qre-value">{d(supplies)}</div></div>'
        f'<div class="qre-card"><div class="qre-label">Computers / Cloud</div><div class="qre-value">{d(comps)}</div></div>'
        f'<div class="qre-card"><div class="qre-label">Contract Research (65%)</div><div class="qre-value">{d(contract)}</div></div>'
        f'<div class="qre-card qre-total"><div class="qre-label">Total QREs</div><div class="qre-value">{d(total)}</div></div>'
        f'</div>\n'
        f'<div class="card-header" style="margin-top:0.5rem">Credit Calculation Method</div>'
        f'<div class="method-row"><span class="method-label">Method Applied:</span><span class="method-value">{esc(method_str)}</span></div>\n'
        f'<div class="credit-banner">'
        f'<div class="credit-banner-label">Estimated Federal R&amp;D Tax Credit</div>'
        f'<div class="{credit_class}">{credit_disp}</div>'
        f'{tax_savings_html}'
        f'</div>\n'
        f'{warn_html}'
    )

    return (
        '<div class="page">'
        + _page_header(fields)
        + '<div class="page-content">' + body + '</div>'
        + _page_footer()
        + '</div>\n'
    )


# ---------------------------------------------------------------------------
# HTML wrapper
# ---------------------------------------------------------------------------

def wrap_html(body_html: str, title: str) -> str:
    return (
        '<!DOCTYPE html>\n<html lang="en">\n<head>\n'
        '<meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
        '<title>' + esc(title) + '</title>\n'
        '<style>\n' + CSS + '\n</style>\n'
        '</head>\n<body>\n'
        + body_html +
        '</body>\n</html>'
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    global LOGO_B64

    script_dir = Path(__file__).parent
    input_dir  = script_dir / "Input"
    output_dir = script_dir / "Output"

    logo_path = input_dir / "logo.png"
    if logo_path.exists():
        LOGO_B64 = "data:image/png;base64," + base64.b64encode(logo_path.read_bytes()).decode()
    else:
        print("  logo.png not found in Input/")

    scenario_pattern = str(output_dir / "rd_credit_calculator_v2_scenario_*.xlsx")
    scenario_paths   = sorted(glob.glob(scenario_pattern))
    if not scenario_paths:
        print("No scenario files found in Output/. Run rd_credit_calculator_v2.py first.")
        sys.exit(1)
    print(f"Found {len(scenario_paths)} scenario file(s).")

    fields = load_company_info(scenario_paths)
    if not fields:
        print("ERROR: Could not read company info from scenario files.")
        sys.exit(1)

    company  = fields.get("company_name", "Company")
    tax_year = fields.get("tax_year", "")
    print(f"Company: {company} | Tax Year: {tax_year}")

    html_body = ""
    html_body += build_cover_pages(fields)
    html_body += build_content_pages(fields)

    for path in scenario_paths:
        print(f"  Adding scenario: {Path(path).name}")
        result = load_scenario(path)
        html_body += build_scenario_page(result, fields)

    safe_company = re.sub(r'[^A-Za-z0-9_]', '_', company)
    out_name = f"rd_feasibility_{safe_company}_{tax_year}.html"
    out_path = output_dir / out_name

    out_path.write_text(wrap_html(html_body, title=f"R&D Tax Credit Feasibility Study - {company} {tax_year}"), encoding="utf-8")
    print(f"\n  HTML written to: {out_path}")

    if HAS_PLAYWRIGHT:
        pdf_path = out_path.with_suffix(".pdf")
        print("  Generating PDF via Playwright...")
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.goto(out_path.as_uri())
            page.wait_for_load_state("networkidle")
            page.pdf(
                path=str(pdf_path),
                format="Letter",
                print_background=True,
                margin={"top": "0", "bottom": "0", "left": "0", "right": "0"},
            )
            browser.close()
        print(f"  PDF written to:  {pdf_path}")
    else:
        print("  (Playwright not found — PDF skipped. Run: pip install playwright && playwright install chromium)")


if __name__ == "__main__":
    main()
