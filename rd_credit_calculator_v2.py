"""
R&D Tax Credit Calculator — Dual Method (Regular Credit + ASC)
IRC Section 41 / Form 6765

Reads from: Input/R&D_Credit_Calculations.xlsx  (sheet: Calculation_Data)
  Column A = field_id
  Column B = value
  Rule: blank = missing data | 0 = actual zero

Outputs: rd_credit_results_v2_<company>_<year>.csv  →  Output/ folder

Usage:
    python rd_credit_calculator_v2.py [path_to_excel]
    (if no path given, auto-detects the xlsx in Input/)
"""

import sys
import os
import csv
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl


# ---------------------------------------------------------------------------
# Constants — IRC §41 / Form 6765
# ---------------------------------------------------------------------------

CONTRACT_US_PCT         = 0.65   # 65% of U.S. contractor costs  §41(b)(3)
REGULAR_RATE            = 0.20   # 20%  regular credit rate
REGULAR_280C_FACTOR     = 0.79   # ×79% with 280C election  → net 15.8%
SPECIAL_RATE            = 0.20   # 20%  energy consortia / qualified orgs
ASC_RATE                = 0.14   # 14%  standard ASC  (line 25)
ASC_STARTUP_RATE        = 0.06   # 6%   startup ASC (any prior-yr QRE = 0)
ASC_280C_FACTOR         = 0.79   # ×79% with 280C  (line 26)
STARTUP_FIXED_BASE_PCT  = 0.03   # 3%   fixed-base % for QRE years 1–5


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_data(excel_path: str) -> dict:
    """Read Calculation_Data tab; return dict keyed by field_id (col A → col B)."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if "Calculation_Data" not in wb.sheetnames:
        raise ValueError(
            f"Sheet 'Calculation_Data' not found. Available: {wb.sheetnames}"
        )

    ws = wb["Calculation_Data"]
    data = {}

    for row in ws.iter_rows(min_row=2, values_only=True):   # skip header row
        field_id = row[0]
        value    = row[1]   # col B — the actual value
        if field_id:
            data[str(field_id).strip()] = value

    return data


# ---------------------------------------------------------------------------
# Helper — value extraction with blank/zero semantics
# ---------------------------------------------------------------------------

def _to_float(v):
    """Convert a cell value to float; return None if blank/None."""
    if v is None:
        return None
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return None


def get_value(data: dict, key: str) -> float | None:
    """
    Return float or None.
      None  → blank / missing (the strict missing-data signal)
      0.0   → explicit zero
    """
    return _to_float(data.get(key))


def get_value_default_zero(data: dict, key: str) -> float:
    """Like get_value but blank → 0.0 (used for optional fields such as qre_computers)."""
    v = get_value(data, key)
    return 0.0 if v is None else v


# ---------------------------------------------------------------------------
# Calculation engine
# ---------------------------------------------------------------------------

def calculate(data: dict) -> dict:
    warnings = []

    # ── Basic identifiers ────────────────────────────────────────────────────
    company_name         = str(data.get("company_name", "Unknown") or "Unknown").strip()
    entity_type          = str(data.get("entity_type",  "Unknown") or "Unknown").strip()
    tax_year_raw         = data.get("tax_year")
    year_research_raw    = data.get("year_research_started")

    tax_year = int(tax_year_raw) if tax_year_raw else None
    year_research_started = int(year_research_raw) if year_research_raw else None

    year_started_raw = data.get("year_started")
    year_started = int(year_started_raw) if year_started_raw else None   # year company was incorporated/started

    scenario_number = int(data.get("scenario_number") or 1)

    # ── Step 1 — current-year QRE components ─────────────────────────────────
    qre_wages            = get_value_default_zero(data, "qre_wages")
    qre_supplies         = get_value_default_zero(data, "qre_supplies")
    qre_computers        = get_value_default_zero(data, "qre_computers")   # blank → 0
    qre_contract_US_raw  = get_value_default_zero(data, "qre_contract_US")
    # qre_contract_foreign is excluded from all credit calculations per §41

    qualified_contract_US = qre_contract_US_raw * CONTRACT_US_PCT

    ordinary_qre = (
        qre_wages
        + qre_supplies
        + qre_computers
        + qualified_contract_US
    )

    # ── Step 2 — special research credit items ────────────────────────────────
    # These receive a direct 20% / 15.8% credit (not part of the base-amount pool)
    energy_consortia      = get_value_default_zero(data, "energy_consortia")
    qualified_orgs        = get_value_default_zero(data, "qualified_organizations")
    special_amount        = energy_consortia + qualified_orgs
    special_credit_before = special_amount * SPECIAL_RATE
    special_credit_280c   = special_amount * SPECIAL_RATE * REGULAR_280C_FACTOR

    # ── Step 3 — determine QRE year number ───────────────────────────────────
    if tax_year and year_research_started:
        qre_year_number = tax_year - year_research_started + 1
    else:
        qre_year_number = None

    # ── Step 4 — Regular Credit Method ───────────────────────────────────────
    regular_credit_280c   = None
    regular_method_status = ""

    if qre_year_number is None:
        regular_method_status = (
            "Insufficient data — year_research_started or tax_year missing"
        )
        warnings.append(f"Regular Method: {regular_method_status}")

    elif 1 <= qre_year_number <= 5:
        fixed_base_pct = STARTUP_FIXED_BASE_PCT          # 3% — no prior QRE history needed for years 1–5

        # Resolve 4 prior-year gross receipts with year_started awareness:
        #   • year before year_started  → company didn't exist, auto-set to 0
        #   • year >= year_started, value is blank → missing data, regular method unavailable
        #   • value is 0 → actual zero, use it
        gr_resolved = []
        missing_gr  = []
        for i in range(1, 5):
            prior_year = tax_year - i if tax_year else None
            raw_val    = get_value(data, f"annual_gross_receipts_yr_minus{i}")
            if prior_year is not None and year_started is not None and prior_year < year_started:
                gr_resolved.append(0.0)   # company didn't exist yet
            elif raw_val is None:
                missing_gr.append(i)      # blank for a year company existed
                gr_resolved.append(None)
            else:
                gr_resolved.append(raw_val)

        if missing_gr:
            regular_method_status = (
                "Regular method not calculated because prior-year gross receipts data "
                "was not fully provided. Blank fields are treated as missing data, not "
                "zero, to avoid overstating the credit. Simplified method is presented "
                "as the feasibility estimate."
            )
            warnings.append(f"Regular Method: gross receipts blank for yr_minus{missing_gr}")
        else:
            avg_gross_receipts = sum(gr_resolved) / len(gr_resolved)
            base_amount        = avg_gross_receipts * fixed_base_pct
            qre_excess         = max(ordinary_qre - base_amount, 0.0)
            fifty_pct_limit    = ordinary_qre * 0.50
            regular_base       = min(qre_excess, fifty_pct_limit)

            regular_credit_raw  = regular_base * REGULAR_RATE            # ×20%
            regular_credit_280c = (
                regular_credit_raw * REGULAR_280C_FACTOR                 # ×79%
                + special_credit_280c
            )
            regular_method_status = (
                f"QRE Year {qre_year_number} — Fixed-base 3% "
                f"(base ${base_amount:,.2f}, excess ${qre_excess:,.2f})"
            )

    elif qre_year_number >= 6:
        # Years 6–10 require historical QRE & gross-receipts data for special formulas
        # Check if sufficient data exists; if not, flag as insufficient
        qre_hist   = [get_value(data, f"qre_yr_minus{i}") for i in range(1, 11)]
        gr_hist    = [get_value(data, f"annual_gross_receipts_yr_minus{i}") for i in range(1, 11)]
        has_data   = any(v is not None for v in qre_hist) and any(v is not None for v in gr_hist)

        if not has_data:
            regular_method_status = (
                f"Insufficient data — QRE year {qre_year_number} requires "
                "historical QRE and gross receipts for fixed-base % computation"
            )
            warnings.append(f"Regular Method: {regular_method_status}")
        else:
            # Year 6–10 special fixed-base formulas (per IRS instructions §41 / Form 6765 Line 6)
            def _calc_pct(qre_yrs, gr_yrs, divisor):
                """Aggregate QRE / aggregate GR / divisor, capped at 16%."""
                total_qre = sum(get_value(data, f"qre_yr_minus{y}") or 0 for y in qre_yrs)
                total_gr  = sum(get_value(data, f"annual_gross_receipts_yr_minus{y}") or 0 for y in gr_yrs)
                if total_gr == 0:
                    return 0.0
                return min((total_qre / total_gr) / divisor, 0.16)

            if qre_year_number == 6:
                fixed_base_pct = _calc_pct([4, 5], [4, 5], 6)
            elif qre_year_number == 7:
                fixed_base_pct = _calc_pct([5, 6], [5, 6], 3)
            elif qre_year_number == 8:
                fixed_base_pct = _calc_pct([5, 6, 7], [5, 6, 7], 2)
            elif qre_year_number == 9:
                fixed_base_pct = _calc_pct([5, 6, 7, 8], [5, 6, 7, 8], 1.5)
            elif qre_year_number == 10:
                fixed_base_pct = _calc_pct([5, 6, 7, 8, 9], [5, 6, 7, 8, 9], 1.2)
            else:   # 11+
                # Use any 5 of QRE years 5–10 (script uses 5–9 by default)
                fixed_base_pct = _calc_pct([5, 6, 7, 8, 9], [5, 6, 7, 8, 9], 1.0)

            # Round to nearest 1/100th of 1% (4 decimal places)
            fixed_base_pct = round(fixed_base_pct, 4)

            gr_4yr = [get_value(data, f"annual_gross_receipts_yr_minus{i}") for i in range(1, 5)]
            missing_gr = [i for i, v in enumerate(gr_4yr, 1) if v is None]

            if missing_gr:
                regular_method_status = (
                    f"Insufficient data — gross receipts missing for yr_minus{missing_gr}"
                )
                warnings.append(f"Regular Method: {regular_method_status}")
            else:
                avg_gross_receipts  = sum(gr_4yr) / len(gr_4yr)
                base_amount         = avg_gross_receipts * fixed_base_pct
                qre_excess          = max(ordinary_qre - base_amount, 0.0)
                fifty_pct_limit     = ordinary_qre * 0.50
                regular_base        = min(qre_excess, fifty_pct_limit)

                regular_credit_raw  = regular_base * REGULAR_RATE
                regular_credit_280c = (
                    regular_credit_raw * REGULAR_280C_FACTOR
                    + special_credit_280c
                )
                regular_method_status = (
                    f"QRE Year {qre_year_number} — Fixed-base {fixed_base_pct*100:.4f}% "
                    f"(base ${base_amount:,.2f}, excess ${qre_excess:,.2f})"
                )

    # ── Step 5 — Alternative Simplified Credit (ASC) ─────────────────────────
    asc_credit_280c   = None
    asc_method_status = ""

    yr1 = get_value(data, "qre_yr_minus1")
    yr2 = get_value(data, "qre_yr_minus2")
    yr3 = get_value(data, "qre_yr_minus3")

    missing_asc = [i for i, v in enumerate([yr1, yr2, yr3], 1) if v is None]

    if missing_asc:
        asc_method_status = (
            f"Insufficient data — prior-year QREs missing for "
            f"yr_minus{missing_asc} (enter 0 if actual zero)"
        )
        warnings.append(f"ASC Method: {asc_method_status}")

    elif yr1 == 0 or yr2 == 0 or yr3 == 0:
        # Startup: any one of the 3 prior years has QRE = 0  →  6% rate
        asc_raw         = ordinary_qre * ASC_STARTUP_RATE          # 6%
        asc_credit_280c = asc_raw * ASC_280C_FACTOR + special_credit_280c
        asc_method_status = "ASC Startup (6%) — prior-year QRE = 0 in at least one year"

    else:
        # Standard ASC: all 3 prior years have QREs > 0
        avg_prior_3yr   = (yr1 + yr2 + yr3) / 3
        excess          = max(ordinary_qre - 0.50 * avg_prior_3yr, 0.0)
        asc_raw         = excess * ASC_RATE                         # 14%
        asc_credit_280c = asc_raw * ASC_280C_FACTOR + special_credit_280c
        asc_method_status = (
            f"ASC Standard (14%) — avg prior 3-yr QRE ${avg_prior_3yr:,.2f}, "
            f"excess ${excess:,.2f}"
        )

    # ── Step 6 — Select recommended credit ───────────────────────────────────
    valid = {}
    if regular_credit_280c is not None:
        valid["Regular Method"] = regular_credit_280c
    if asc_credit_280c is not None:
        valid["ASC Method"]     = asc_credit_280c

    if valid:
        best_method       = max(valid, key=lambda k: valid[k])
        recommended_credit = valid[best_method]
        recommended_note   = f"{best_method} selected (higher 280C-reduced credit)"
    else:
        recommended_credit = None
        recommended_note   = "N/A — both methods have insufficient data"

    return {
        # Identifiers
        "company_name":            company_name,
        "tax_year":                tax_year,
        "entity_type":             entity_type,
        "electing_280c":           "Yes",
        # QRE components
        "qre_wages":               qre_wages,
        "qre_supplies":            qre_supplies,
        "qre_computers":           qre_computers,
        "qualified_contract_US":   qualified_contract_US,
        "ordinary_qre":            ordinary_qre,
        "fifty_pct_qre":           ordinary_qre * 0.50,
        # Method results (with 280C)
        "regular_credit_280c":     regular_credit_280c,
        "regular_method_status":   regular_method_status,
        "asc_credit_280c":         asc_credit_280c,
        "asc_method_status":       asc_method_status,
        # Recommendation
        "recommended_credit":      recommended_credit,
        "recommended_note":        recommended_note,
        # Warnings
        "warnings":                warnings,
        # Detail (for console summary)
        "qre_year_number":         qre_year_number,
        "scenario_number":         scenario_number,
    }


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def fmt(v):
    """Format a number as a dollar string, or return the string as-is."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return "$ -" if v == 0 else f"$ {v:,.2f}"
    return str(v)


def credit_cell(credit_value, status_str):
    """Return dollar amount if available, otherwise the status/error message."""
    if credit_value is not None:
        return fmt(credit_value)
    return status_str


def save_csv(result: dict, output_path: str):
    """Write 13-row results CSV."""
    rows = [
        ("Label",                                                              "Value"),
        ("Company Name",                                                       result["company_name"]),
        ("Tax Year",                                                           result["tax_year"]),
        ("Entity Type",                                                        result["entity_type"]),
        ("Are you electing the reduced credit under section 280C?",            result["electing_280c"]),
        ("Wages for qualified services",                                       fmt(result["qre_wages"])),
        ("Cost of supplies",                                                   fmt(result["qre_supplies"])),
        ("Rental or lease costs of computers",                                 fmt(result["qre_computers"])),
        ("Applicable percentage of contract research expenses (65%)",          fmt(result["qualified_contract_US"])),
        ("Total qualified research expenses",                                  fmt(result["ordinary_qre"])),
        ("50% of total qualified research expenses",                           fmt(result["fifty_pct_qre"])),
        ("Regular Credit Method — 280C applied (15.8%)",                      credit_cell(result["regular_credit_280c"], result["regular_method_status"])),
        ("Alternative Simplified Credit (ASC) — 280C applied",                credit_cell(result["asc_credit_280c"],     result["asc_method_status"])),
        ("Recommended R&D Tax Credit",                                         credit_cell(result["recommended_credit"],  result["recommended_note"])),
    ]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)

    print(f"\n✓ Results saved to: {output_path}")


def save_xlsx(result: dict, raw_data: dict, output_path: str):
    """
    Write results to Excel.
    Sheet 1 — Input Data: all raw input fields that have a value (non-blank).
    Sheet 2 — Calculated Results: the 13-row results table.
    """
    wb = openpyxl.Workbook()

    # ── Styles ──────────────────────────────────────────────────────────────
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="003366")
    section_font = Font(bold=True, color="003366")
    dollar_fmt   = '#,##0.00'

    def _header_row(ws, col_a, col_b, col_a_width=40, col_b_width=30):
        row = ws.max_row + 1
        for col, val in [(1, col_a), (2, col_b)]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions["A"].width = col_a_width
        ws.column_dimensions["B"].width = col_b_width

    def _data_row(ws, label, value):
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=label)
        if isinstance(value, (int, float)):
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = dollar_fmt
        else:
            ws.cell(row=row, column=2, value=value)

    def _blank_row(ws):
        ws.append([])

    # ── Sheet 1: Input Data ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Input Data"
    _header_row(ws1, "Field", "Value")

    # Friendly label map for known fields
    labels = {
        "company_name":            "Company Name",
        "entity_type":             "Entity Type",
        "tax_year":                "Tax Year",
        "EIN_number":              "EIN",
        "study_date":              "Study Date",
        "preparer_name":           "Preparer Name",
        "year_started":            "Year Company Started",
        "year_research_started":   "Year Research Started",
        "scenario_number":         "Scenario Number",
        "qre_wages":               "Wages for Qualified Services",
        "qre_supplies":            "Cost of Supplies",
        "qre_computers":           "Rental / Lease Cost of Computers",
        "qre_contract_US":         "U.S. Contract Research (gross)",
        "qre_contract_foreign":    "Foreign Contract Research (gross)",
        "energy_consortia":        "Energy Consortia Payments",
        "qualified_organizations": "Qualified Organization Payments",
        "basic_research_payments": "Basic Research Payments",
        "company_industry":        "Industry",
    }
    # Prior-year QRE labels
    for i in range(1, 16):
        labels[f"qre_yr_minus{i}"] = f"Prior QRE — Year Minus {i}"
    # Gross receipts labels
    for i in range(1, 16):
        labels[f"annual_gross_receipts_yr_minus{i}"] = f"Annual Gross Receipts — Year Minus {i}"

    for field_id, value in raw_data.items():
        if value is None:
            continue   # skip blank fields
        label = labels.get(field_id, field_id)
        # Format datetime objects
        if hasattr(value, "strftime"):
            value = value.strftime("%Y-%m-%d")
        _data_row(ws1, label, value)

    # ── Sheet 2: Calculated Results ─────────────────────
    ws2 = wb.create_sheet("Calculated Results")
    _header_row(ws2, "Label", "Value")

    def _result_row(ws, label, value, status=None):
        if value is not None:
            _data_row(ws, label, value)
        else:
            _data_row(ws, label, status or "")

    _data_row(ws2, "Company Name",                                            result["company_name"])
    _data_row(ws2, "Tax Year",                                                result["tax_year"])
    _data_row(ws2, "Scenario Number",                                         result["scenario_number"])
    _data_row(ws2, "Entity Type",                                             result["entity_type"])
    _data_row(ws2, "Are you electing the reduced credit under section 280C?", result["electing_280c"])
    _blank_row(ws2)
    _data_row(ws2, "Wages for qualified services",                            result["qre_wages"])
    _data_row(ws2, "Cost of supplies",                                        result["qre_supplies"])
    _data_row(ws2, "Rental or lease costs of computers",                      result["qre_computers"])
    _data_row(ws2, "Applicable % of contract research expenses (65%)",        result["qualified_contract_US"])
    _data_row(ws2, "Total qualified research expenses",                       result["ordinary_qre"])
    _data_row(ws2, "50% of total qualified research expenses",                result["fifty_pct_qre"])
    _blank_row(ws2)
    _result_row(ws2, "Regular Credit Method — 280C applied (15.8%)",
                result["regular_credit_280c"],  result["regular_method_status"])
    _result_row(ws2, "Alternative Simplified Credit (ASC) — 280C applied",
                result["asc_credit_280c"],      result["asc_method_status"])
    _blank_row(ws2)
    _result_row(ws2, "Recommended R&D Tax Credit",
                result["recommended_credit"],   result["recommended_note"])

    if result["warnings"]:
        _blank_row(ws2)
        cell = ws2.cell(row=ws2.max_row + 1, column=1, value="Warnings / Notes")
        cell.font = section_font
        for msg in result["warnings"]:
            _data_row(ws2, "", msg)

    wb.save(output_path)
    print(f"\n✓ Results saved to: {output_path}")


def print_summary(result: dict):
    w = result["warnings"]
    print("\n" + "=" * 65)
    print(f"  R&D TAX CREDIT — {result['company_name']} ({result['tax_year']})")
    print("=" * 65)
    print(f"  Entity Type        : {result['entity_type']}")
    print(f"  280C Election      : {result['electing_280c']}")
    print(f"  QRE Year #         : {result['qre_year_number'] or 'N/A'}")
    print(f"  Scenario #         : {result['scenario_number']}")
    print(f"  " + "─" * 61)
    print(f"  Wages              : ${result['qre_wages']:>18,.2f}")
    print(f"  Supplies           : ${result['qre_supplies']:>18,.2f}")
    print(f"  Computers          : ${result['qre_computers']:>18,.2f}")
    print(f"  Contract US (65%)  : ${result['qualified_contract_US']:>18,.2f}")
    print(f"  " + "─" * 61)
    print(f"  Total QREs         : ${result['ordinary_qre']:>18,.2f}")
    print(f"  50% of QREs        : ${result['fifty_pct_qre']:>18,.2f}")
    print(f"  " + "─" * 61)

    if result["regular_credit_280c"] is not None:
        print(f"  Regular Credit     : ${result['regular_credit_280c']:>18,.2f}")
    else:
        print(f"  Regular Credit     : {result['regular_method_status']}")

    if result["asc_credit_280c"] is not None:
        print(f"  ASC Credit         : ${result['asc_credit_280c']:>18,.2f}")
    else:
        print(f"  ASC Credit         : {result['asc_method_status']}")

    print(f"  " + "─" * 61)
    if result["recommended_credit"] is not None:
        print(f"  RECOMMENDED CREDIT : ${result['recommended_credit']:>18,.2f}  ← {result['recommended_note']}")
    else:
        print(f"  RECOMMENDED CREDIT : {result['recommended_note']}")

    if w:
        print(f"\n  Warnings:")
        for msg in w:
            print(f"    ⚠  {msg}")
    print("=" * 65)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_dir  = os.path.join(script_dir, "Input")
    output_dir = os.path.join(script_dir, "Output")

    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        default_file = os.path.join(input_dir, "R&D_Credit_Calculations.xlsx")
        if not os.path.exists(default_file):
            print(f"ERROR: Expected input file not found: {default_file}")
            sys.exit(1)
        excel_path = default_file

    print(f"Reading: {excel_path}")

    data   = load_data(excel_path)
    result = calculate(data)

    scenario  = result["scenario_number"]
    filename  = f"rd_credit_calculator_v2_scenario_{scenario}.xlsx"
    out_path  = os.path.join(output_dir, filename)

    save_xlsx(result, data, out_path)
    print_summary(result)


if __name__ == "__main__":
    main()
