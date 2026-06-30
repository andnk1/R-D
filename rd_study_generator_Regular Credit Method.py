#!/usr/bin/env python3
"""
rd_study_generator.py
Reads  Input/R&D_Credit_Questionnaire.xlsx
Writes Output/rd_study_<Company>_<Year>.html

Requirements: openpyxl   (pip install openpyxl)

Cover letter is generated from the LETTER_PARAS template defined in this file.
Edit LETTER_PARAS to customize the standard letter for all clients.
"""
import re, sys, base64
from pathlib import Path
from datetime import datetime

LOGO_B64 = ""   # populated at runtime from Input/logo.png

LETTER_PARAS = [
    "[study_date]",
    "EIN: [EIN_number]",
    "Re: [Company_Name] [tax_year] Research & Development Tax Credit Results",
    "Tax year [tax_year]",
    "",
    "Enclosed is the R&D tax credit calculation we have prepared for your review. This report provides detailed documentation of the research activities and related expenses undertaken by [Company_Name] for the [tax_year] in support of our claim for the Research and Development (R&D) Tax Credit under Section 41. This report complies with IRS requirements for filing a valid claim for refund.",
    "",
    "Table of Contents",
    "",
    "1. Credit for Increasing Research Activities Summary",
    "2. Company Overview",
    "3. Project Overview",

    "4. Wage QRE",
    "5. Supply QRE",
    "6. Contractor US",
    "7. Contractor Foreign",
]

CREDIT_TAB_NARRATIVE = [
    "For the tax year [tax_year], [company_name], a [entity_type], completed a review of its qualified research activities and related expenditures for purposes of computing the Credit for Increasing Research Activities under IRC Section 41.",
    "The calculation was prepared based on qualified research expenditures such as wages for qualified services of [qre_wages], supplies of [qre_supplies], U.S. contract research of [qre_contract_US], foreign contract research of [qre_contract_foreign], and basic research payments of [basic_research_payments]. Supporting details for the calculations are presented in the text sections. The calculation reflects the Section 280C reduced credit election, as shown on the table above.",
    "Based on the analysis performed, [company_name] computed an estimated federal R&D tax credit of [rd_credit_amount] for the [tax_year] tax year.",
    "This study was prepared on [study_date] by CFO Associates and is intended to summarize the qualified cost categories, methodology, and assumptions used in the credit calculation. The final credit should be supported by company books, payroll records, project documentation, and other relevant business records.",
]

# Dollar fields — these get formatted as $X,XXX.XX when used in narrative wildcards
DOLLAR_FIELDS = {
    "qre_wages", "qre_supplies", "qre_contract_us", "qre_contract_foreign",
    "basic_research_payments", "rd_credit_amount",
}

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _cv(v):
    if v is None: return ""
    return str(v).strip()

def _num(v):
    if v is None or v == "": return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace("$","").replace(",","").replace("\t","").strip()
    try: return float(s)
    except ValueError: return 0.0

def fmt_dollar(v):
    n = _num(v)
    return "$ -" if n == 0 else f"${n:,.2f}"

def fmt_dollar_total(v):
    return f"${_num(v):,.2f}"

def fmt_pct(v):
    if v is None or v == "": return ""
    try:
        n = float(v) if not isinstance(v, str) else float(str(v).replace("%","").strip())
        if n <= 1.0 and "%" not in str(v): n *= 100
        return f"{int(n)}%" if n == int(n) else f"{n:.1f}%"
    except: return str(v)

def fmt_date(v):
    if v is None or v == "": return ""
    if isinstance(v, datetime): return v.strftime("%B %d, %Y")
    return str(v)

def esc(text):
    return (str(text).replace("&","&amp;").replace("<","&lt;")
            .replace(">","&gt;").replace('"',"&quot;"))

def apply_wildcards(text, fields):
    def replacer(m):
        key = m.group(1).lower()
        val = fields.get(key)
        if val is None: return m.group(0)
        if isinstance(val, datetime): return val.strftime("%B %d, %Y")
        if key in DOLLAR_FIELDS:
            try: return f"${float(val):,.2f}"
            except: return str(val)
        if isinstance(val, float):
            if val == int(val): return str(int(val))
            return str(val)
        if isinstance(val, int): return str(val)
        return str(val)
    return re.sub(r'\[([A-Za-z_]+)\]', replacer, text)

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL READERS
# ═══════════════════════════════════════════════════════════════════════════════
def read_calc_data(ws):
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fid = _cv(row[0]) if len(row) > 0 else ""
        if not fid: continue
        val = row[2] if len(row) > 2 else None
        data[fid.lower()] = val
    return data

def read_qa_sheet(ws):
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        q = _cv(row[0]) if len(row) > 0 else ""
        a = row[1]       if len(row) > 1 else None
        if q:
            if isinstance(a, datetime): a = fmt_date(a)
            rows.append((q, _cv(a)))
    return rows

def read_wage_qre(ws):
    total_qre = 0.0; total_taxable = 0.0; data_rows = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if ri == 0:
            if len(row) > 7 and _cv(row[6]).lower() == "calculated_qre_wages":
                total_qre = _num(row[7])
            continue
        name = _cv(row[0]) if len(row) > 0 else ""
        if not name: continue
        title   = _cv(row[1]) if len(row) > 1 else ""
        state   = _cv(row[2]) if len(row) > 2 else ""
        taxable = _num(row[3]) if len(row) > 3 else 0.0
        qual    = row[4]       if len(row) > 4 else None
        qre_amt = _num(row[5]) if len(row) > 5 else 0.0
        total_taxable += taxable
        data_rows.append((name, title, state, taxable, qual, qre_amt))
    if total_qre == 0.0 and data_rows:
        total_qre = sum(r[5] for r in data_rows)
    return data_rows, total_taxable, total_qre

def read_supply_qre(ws):
    total_qre = 0.0; data_rows = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if len(row) > 8 and _cv(row[7]).lower() == "qre_supplies":
            total_qre = _num(row[8]); continue
        if ri == 0: continue
        vendor = _cv(row[0]) if len(row) > 0 else ""
        if not vendor: continue
        state  = _cv(row[1]) if len(row) > 1 else ""
        type_  = _cv(row[2]) if len(row) > 2 else ""
        amount = _num(row[3]) if len(row) > 3 else 0.0
        qual   = row[4]       if len(row) > 4 else None
        total  = _num(row[5]) if len(row) > 5 else 0.0
        data_rows.append((vendor, state, type_, amount, qual, total))
    if total_qre == 0.0 and data_rows:
        total_qre = sum(r[5] for r in data_rows)
    return data_rows, total_qre

def read_contractor_sheet(ws):
    grand_total = 0.0; data_rows = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if ri == 0:
            if len(row) > 7: grand_total = _num(row[7])
            continue
        name = _cv(row[0]) if len(row) > 0 else ""
        if not name: continue
        loc    = _cv(row[1]) if len(row) > 1 else ""
        amount = _num(row[2]) if len(row) > 2 else 0.0
        qual   = row[3]       if len(row) > 3 else None
        work   = _cv(row[4])  if len(row) > 4 else ""
        total  = _num(row[5]) if len(row) > 5 else 0.0
        data_rows.append((name, loc, amount, qual, work, total))
    if grand_total == 0.0 and data_rows:
        grand_total = sum(r[5] for r in data_rows)
    return data_rows, grand_total

def read_gross_receipts(ws, tax_year_int):
    yr_m2 = str(tax_year_int - 2); yr_m1 = str(tax_year_int - 1)
    col_headers = ["Income Item", yr_m2, yr_m1]
    data_rows = []; total_row = None
    def safe_num(x):
        if x in (None, "", "$ -", "$\t-"): return 0.0
        return _num(x)
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if ri == 0: continue
        item = _cv(row[0]) if len(row) > 0 else ""
        if not item: continue
        v2 = safe_num(row[1] if len(row) > 1 else None)
        v3 = safe_num(row[2] if len(row) > 2 else None)
        if "total" in item.lower():
            total_row = (item, v2, v3)
        else:
            data_rows.append((item, v2, v3))
    return col_headers, data_rows, total_row

def read_comments(ws):
    rows = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if ri == 0: continue
        ctype   = _cv(row[0]) if len(row) > 0 else ""
        comment = _cv(row[1]) if len(row) > 1 else ""
        include = _cv(row[2]).strip().lower() if len(row) > 2 else ""
        if include == "yes" and comment:
            rows.append((ctype, comment))
    return rows

def read_text_file(path):
    """Read a .txt narrative file and return a list of paragraph strings.
    Blank lines become paragraph breaks. Supports [field] wildcards."""
    try:
        text = Path(path).read_text(encoding="utf-8")
    except FileNotFoundError:
        return []
    # Split on blank lines to get paragraphs; strip each line within a paragraph
    paragraphs = []
    for block in re.split(r'\n\s*\n', text.strip()):
        para = " ".join(line.strip() for line in block.splitlines() if line.strip())
        if para:
            paragraphs.append(para)
    return paragraphs

def read_credit_tab(ws):
    """Read the R&D TAX CREDIT tab — two columns: Label, Value.
    Numeric values are formatted as dollar amounts; text values are kept as-is."""
    rows = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if ri == 0: continue          # skip header
        label = _cv(row[0]) if len(row) > 0 else ""
        raw   = row[1]       if len(row) > 1 else None
        if isinstance(raw, (int, float)):
            # Year values (e.g. 2023) — plain integer, no $ or comma
            if isinstance(raw, (int, float)) and 1900 <= raw <= 2100:
                value = str(int(raw))
            elif raw == 0:
                value = "$ -"
            else:
                value = f"$ {raw:,.2f}"
        else:
            value = _cv(raw)
        if label:
            rows.append((label, value))
    return rows

# ═══════════════════════════════════════════════════════════════════════════════
# COVER LETTER  — built from LETTER_PARAS template above; edit that list to
#                 customise the standard letter for all clients.
# ═══════════════════════════════════════════════════════════════════════════════
def build_letter_paras(fields):
    """Apply field wildcards to LETTER_PARAS and return resolved paragraph list."""
    result = []
    for txt in LETTER_PARAS:
        result.append(apply_wildcards(txt, fields))
    return result

# ═══════════════════════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════════════════════
CSS = """
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

body {
    font-family: Verdana, Geneva, Tahoma, sans-serif;
    font-size: 9pt;
    color: #1a1a1a;
    background: #e8e8e8;
    padding: 1.2rem 0 2rem 0;
}

.page {
    width: 8.5in;
    min-height: 11in;
    margin: 0 auto 1.2rem auto;
    background: #fff;
    box-shadow: 0 3px 16px rgba(0,0,0,.18);
    display: flex;
    flex-direction: column;
    padding: 0;
}

@media print {
    body { background: #fff; padding: 0; }
    .page {
        box-shadow: none;
        margin: 0;
        page-break-after: always;
        page-break-inside: avoid;
    }
    .page:last-child { page-break-after: avoid; }
}

.page-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0.35rem 0.9in;
    background: #003366;
    color: #fff;
    font-size: 7.5pt;
    font-weight: bold;
    letter-spacing: 0.3px;
    flex-shrink: 0;
}
.page-header span { flex: 1; }
.page-header span:nth-child(2) { text-align: center; }
.page-header span:last-child   { text-align: right; }

.page-content {
    flex: 1;
    padding: 0.55in 0.9in 0.4in 0.9in;
    overflow: hidden;
}

.page-footer {
    flex-shrink: 0;
    padding: 0 0.9in 0.3in 0.9in;
}
.footer-rule {
    border: none;
    border-top: 1px solid #b8c8d8;
    margin-bottom: 0.25rem;
}
.footer-text {
    font-size: 7pt;
    color: #888;
    text-align: center;
    letter-spacing: 0.5px;
}

.cover-content {
    flex: 1;
    padding: 0.6in 0.9in 0.4in 0.9in;
    display: flex;
    flex-direction: column;
}
.cover-logo {
    height: 264px;
    margin-bottom: 0.45in;
}
.cover-title-block {
    text-align: center;
    border-bottom: 3px solid #003366;
    padding-bottom: 1.2rem;
    margin-bottom: 1.2rem;
}
.cover-title-block h1 {
    font-family: Verdana, sans-serif;
    font-size: 18pt;
    font-weight: bold;
    color: #003366;
    letter-spacing: 0.5px;
    line-height: 1.2;
}
.cover-title-block h2 {
    font-size: 40pt;
    font-weight: bold;
    color: #003366;
    margin-top: 0.3rem;
}
.cover-title-block h3 {
    font-size: 9.5pt;
    color: #555;
    margin-top: 0.25rem;
}
.cover-title-block .completed-by {
    font-size: 8.5pt;
    color: #555;
    margin-top: 0.4rem;
}
.cover-bottom-info {
    margin-top: auto;
    text-align: right;
    font-size: 9pt;
    color: #555;
    line-height: 1.8;
}
.cover-letter { flex: 1; }
.letter-date { font-size: 9pt; margin: 0.4rem 0; }
.letter-re   { font-size: 9pt; font-weight: bold; margin: 0.5rem 0 0.7rem; }
.cover-letter p {
    font-size: 9pt;
    line-height: 1.65;
    margin-bottom: 0.6rem;
    text-align: justify;
    color: #1a1a1a;
}

.section-title {
    font-family: Verdana, sans-serif;
    font-size: 13pt;
    font-weight: bold;
    color: #003366;
    margin-bottom: 0.1rem;
    padding-bottom: 0.3rem;
    border-bottom: 2px solid #003366;
}

.no-data {
    font-style: italic;
    color: #666;
    margin-top: 0.6rem;
    font-size: 8.5pt;
}

.section-narrative {
    margin-top: 0.55rem;
    font-size: 8.5pt;
    line-height: 1.65;
    color: #1a1a1a;
    text-align: justify;
}
.section-narrative p {
    margin-bottom: 0.45rem;
}

table {
    width: 100%;
    border-collapse: collapse;
    font-family: Verdana, sans-serif;
    font-size: 8pt;
    margin-top: 0.5rem;
}
thead tr { background-color: #003366; color: #fff; }
th {
    text-align: left;
    padding: 0.32rem 0.5rem;
    border: 1px solid #003366;
    font-weight: bold;
    font-size: 7.5pt;
    letter-spacing: 0.2px;
}
td {
    padding: 0.27rem 0.5rem;
    border: 1px solid #ccd8e8;
    vertical-align: top;
}
tr:nth-child(even) td { background-color: #f4f7fb; }
.total-row td {
    background-color: #dbe4f0 !important;
    border-top: 2px solid #003366;
    border-bottom: 2px solid #003366;
    font-weight: bold;
}
.num { text-align: right; font-variant-numeric: tabular-nums; }
"""

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE STRUCTURE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _page_footer():
    return ('<div class="page-footer">'
            '<hr class="footer-rule">'
            '<div class="footer-text">CFO Associates</div>'
            '</div>')

def _page_header(fields):
    company  = esc(fields.get("company_name", ""))
    tax_year = esc(str(fields.get("tax_year", "")))
    return ('<div class="page-header">'
            '<span>' + company + '</span>'
            '<span>Research and Development Credit Study</span>'
            '<span>Tax Year ' + tax_year + '</span>'
            '</div>')

def section_page(title, body_html, fields):
    return ('<div class="page">'
            + _page_header(fields)
            + '<div class="page-content">'
            + '<div class="section-title">' + esc(title) + '</div>'
            + body_html
            + '</div>'
            + _page_footer()
            + '</div>\n')

# ═══════════════════════════════════════════════════════════════════════════════
# COVER PAGE BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
def build_cover_pages(fields):
    company  = esc(fields.get("company_name", ""))
    tax_year = esc(str(fields.get("tax_year", "")))
    preparer = esc(str(fields.get("preparer_name", "")))

    # Page 1 — logo + title block
    page1 = (
        '<div class="page">'
        '<div class="cover-content">'
        '<img src="' + LOGO_B64 + '" class="cover-logo" alt="CFO Associates">'
        '<div class="cover-title-block">'
        '<h1>Research and Development<br>Tax Credit</h1>'
        '<h2>' + company + '</h2>'
        '</div>'
        '<div class="cover-bottom-info">'
        '<div>Tax Year Ended December 31, ' + tax_year + '</div>'
        '<div>Completed by ' + preparer + '</div>'
        '</div>'
        '</div>'
        + _page_footer()
        + '</div>\n'
    )

    # Page 2 (and overflow pages if needed) — cover letter from LETTER_PARAS
    letter_paras = build_letter_paras(fields)
    letter_html = '<div class="cover-content"><div class="cover-letter">'
    for idx, para in enumerate(letter_paras):
        e = esc(para)
        if para.startswith("Re:"):
            letter_html += '<p class="letter-re">' + e + '</p>\n'
        elif para.startswith("•") or para.startswith("*") or para.startswith("-"):
            letter_html += '<p>' + e + '</p>\n'
        elif idx == 0:
            letter_html += '<p class="letter-date">' + e + '</p>\n'
        else:
            letter_html += '<p>' + e + '</p>\n'
    letter_html += '</div></div>'

    page2 = '<div class="page">\n' + letter_html + _page_footer() + '</div>\n'
    return page1 + page2

# ═══════════════════════════════════════════════════════════════════════════════
# SECTION PAGE BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════
def build_credit_tab_page(credit_rows, fields):
    """Page: Credit for Increasing Research Activities — table + narrative."""
    if not credit_rows:
        body = '<p class="no-data">No Data Provided.</p>'
    else:
        body = '<table><tbody>'
        for label, value in credit_rows:
            body += ('<tr><td style="width:75%">' + esc(label) + '</td>'
                     '<td>' + esc(value) + '</td></tr>')
        body += '</tbody></table>'

    # Narrative paragraphs
    body += '<div class="section-narrative">'
    for para in CREDIT_TAB_NARRATIVE:
        body += '<p>' + esc(apply_wildcards(para, fields)) + '</p>'
    body += '</div>'

    return section_page("Credit for Increasing Research Activities", body, fields)

def build_narrative_page(title, paragraphs, fields):
    """Render a section page with free-form text paragraphs (no table)."""
    if not paragraphs:
        return section_page(title, '<p class="no-data">No narrative file found — add '
                            + esc(title.lower().replace(" ", "_") + ".txt")
                            + ' to the Input/ folder.</p>', fields)
    html = '<div class="section-narrative">'
    for para in paragraphs:
        html += '<p>' + esc(apply_wildcards(para, fields)) + '</p>'
    html += '</div>'
    return section_page(title, html, fields)

def build_qa_page(title, qa_rows, fields):
    if not qa_rows:
        return section_page(title, '<p class="no-data">No Data Provided.</p>', fields)
    html = '<table><tbody>'
    for q, a in qa_rows:
        html += '<tr><td style="width:62%">' + esc(q) + '</td><td>' + esc(a) + '</td></tr>'
    html += '</tbody></table>'
    return section_page(title, html, fields)

def build_wage_page(data_rows, total_taxable, total_qre, fields):
    if not data_rows:
        return section_page("Wage QRE", '<p class="no-data">No Data Provided.</p>', fields)
    html = ('<table><thead><tr>'
            '<th>Employee Name</th><th>Job Title</th><th>State</th>'
            '<th class="num">Taxable Wages</th>'
            '<th class="num">Qualified %</th>'
            '<th class="num">QRE Amount</th>'
            '</tr></thead><tbody>')
    for (name, title, state, taxable, qual, qre_amt) in data_rows:
        pct = fmt_pct(qual)
        html += ('<tr><td>' + esc(name) + '</td><td>' + esc(title) + '</td><td>' + esc(state) + '</td>'
                 '<td class="num">' + fmt_dollar(taxable) + '</td>'
                 '<td class="num">' + esc(pct) + '</td>'
                 '<td class="num">' + fmt_dollar(qre_amt) + '</td></tr>')
    html += ('<tr class="total-row"><td colspan="3">Total</td>'
             '<td class="num">' + fmt_dollar_total(total_taxable) + '</td><td></td>'
             '<td class="num">' + fmt_dollar_total(total_qre) + '</td></tr>'
             '</tbody></table>')
    return section_page("Wage QRE", html, fields)

def build_supply_page(data_rows, total_qre, fields):
    if not data_rows:
        return section_page("Supply QRE", '<p class="no-data">No Data Provided.</p>', fields)
    html = ('<table><thead><tr>'
            '<th>Vendor</th><th>State</th><th>Type</th>'
            '<th class="num">Amount</th><th class="num">Qualified %</th>'
            '<th class="num">Qualified Amount</th>'
            '</tr></thead><tbody>')
    for (vendor, state, type_, amount, qual, total) in data_rows:
        pct = fmt_pct(qual)
        html += ('<tr><td>' + esc(vendor) + '</td><td>' + esc(state) + '</td><td>' + esc(type_) + '</td>'
                 '<td class="num">' + fmt_dollar(amount) + '</td>'
                 '<td class="num">' + esc(pct) + '</td>'
                 '<td class="num">' + fmt_dollar(total) + '</td></tr>')
    html += ('<tr class="total-row"><td colspan="5">Total</td>'
             '<td class="num">' + fmt_dollar_total(total_qre) + '</td></tr>'
             '</tbody></table>')
    return section_page("Supply QRE", html, fields)

def build_contractor_page(title, data_rows, grand_total, loc_label, fields):
    if not data_rows:
        return section_page(title, '<p class="no-data">No Data Provided.</p>', fields)
    html = ('<table><thead><tr>'
            '<th>Contractor Name</th><th>' + loc_label + '</th>'
            '<th class="num">Amount</th><th class="num">Qualified %</th>'
            '<th>Work Performed</th><th class="num">Total</th>'
            '</tr></thead><tbody>')
    for (name, loc, amount, qual, work, total) in data_rows:
        pct = fmt_pct(qual)
        html += ('<tr><td>' + esc(name) + '</td><td>' + esc(loc) + '</td>'
                 '<td class="num">' + fmt_dollar(amount) + '</td>'
                 '<td class="num">' + esc(pct) + '</td>'
                 '<td>' + esc(work) + '</td>'
                 '<td class="num">' + fmt_dollar(total) + '</td></tr>')
    html += ('<tr class="total-row"><td colspan="5">Total</td>'
             '<td class="num">' + fmt_dollar_total(grand_total) + '</td></tr>'
             '</tbody></table>')
    return section_page(title, html, fields)

def build_receipts_page(col_headers, data_rows, total_row, fields):
    if not data_rows and not total_row:
        return section_page("Gross Receipts", '<p class="no-data">No Data Provided.</p>', fields)
    h0, h1, h2 = (esc(c) for c in col_headers[:3])
    html = ('<table><thead><tr>'
            '<th>' + h0 + '</th><th class="num">' + h1 + '</th><th class="num">' + h2 + '</th>'
            '</tr></thead><tbody>')
    for (item, v2, v3) in data_rows:
        html += ('<tr><td>' + esc(item) + '</td>'
                 '<td class="num">' + fmt_dollar(v2) + '</td>'
                 '<td class="num">' + fmt_dollar(v3) + '</td></tr>')
    if total_row:
        item, v2, v3 = total_row
        html += ('<tr class="total-row"><td>' + esc(item) + '</td>'
                 '<td class="num">' + fmt_dollar_total(v2) + '</td>'
                 '<td class="num">' + fmt_dollar_total(v3) + '</td></tr>')
    html += '</tbody></table>'
    return section_page("Gross Receipts", html, fields)

def build_comments_page(comment_rows, fields):
    if not comment_rows:
        return section_page("Comments", '<p class="no-data">No Data Provided.</p>', fields)
    html = ('<table><thead><tr>'
            '<th>Type</th><th>Comment</th>'
            '</tr></thead><tbody>')
    for (ctype, comment) in comment_rows:
        html += ('<tr><td style="width:20%">' + esc(ctype) + '</td>'
                 '<td>' + esc(comment) + '</td></tr>')
    html += '</tbody></table>'
    return section_page("Comments", html, fields)

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    global LOGO_B64

    script_dir = Path(__file__).parent
    output_dir = script_dir / "Output"
    output_dir.mkdir(exist_ok=True)

    # Load logo
    input_dir = script_dir / "Input"
    logo_path = input_dir / "logo.png"
    if logo_path.exists():
        LOGO_B64 = "data:image/png;base64," + base64.b64encode(logo_path.read_bytes()).decode()
        print(f"Logo loaded: {logo_path.name}")
    else:
        print("WARNING: Input/logo.png not found — cover logo will be blank.")

    # Locate xlsx
    search_dirs = [input_dir, script_dir] if input_dir.is_dir() else [script_dir]
    xl_path = None
    for d in search_dirs:
        xl_files = [f for f in d.glob("*.xlsx") if not f.name.startswith("~$")]
        if xl_files:
            xl_path = xl_files[0]; break
    if not xl_path:
        print("ERROR: No .xlsx file found in Input/ or next to the script.")
        return
    print(f"Using workbook: {xl_path.name}")

    # Read Excel
    import openpyxl
    wb = openpyxl.load_workbook(xl_path, data_only=True)
    sheets = {s.title.strip().lower(): s for s in wb.worksheets}

    def get_sheet(*names):
        for n in names:
            if n.lower() in sheets:
                return sheets[n.lower()]
        return None

    calc_ws             = get_sheet("Calculation_Data", "Calc Data", "Calculation Data", "Fields", "Data")
    company_overview_ws = get_sheet("Company Overview", "QA", "Q&A", "Taxpayer", "Overview")
    project_ws          = get_sheet("Project Overview", "Project")
    wage_ws             = get_sheet("Wage QRE", "Wages", "Employee Wages")
    supply_ws           = get_sheet("Supply QRE", "Supplies", "Research Supplies")
    contractor_us_ws    = get_sheet("Contractor US", "Contractors", "Contractor QRE", "Contract Research")
    contractor_for_ws   = get_sheet("Contractor Foreign", "Foreign Contractors", "Foreign")
    receipts_ws         = get_sheet("Gross Receipts", "Receipts", "GrossReceipts")
    comments_ws         = get_sheet("Comments", "Notes", "Notes & Comments")
    credit_tab_ws       = get_sheet("R&D TAX CREDIT", "RD TAX CREDIT", "Credit")

    # Fields dict
    fields = {}
    if calc_ws:
        fields = read_calc_data(calc_ws)
    else:
        print("WARNING: Calc Data sheet not found.")

    # Set study_date if missing
    if not fields.get("study_date"):
        fields["study_date"] = datetime.today().strftime("%B %d, %Y")

    # Auto-compute rd_credit_amount if missing
    if not fields.get("rd_credit_amount"):
        _qw  = float(fields.get("qre_wages")      or 0)
        _qs  = float(fields.get("qre_supplies")    or 0)
        _qus = float(fields.get("qre_contract_us") or 0)
        _y1  = float(fields.get("qre_yr_minus1")   or 0)
        _y2  = float(fields.get("qre_yr_minus2")   or 0)
        _y3  = float(fields.get("qre_yr_minus3")   or 0)
        _cq  = _qw + _qs + _qus * 0.65
        _avg = (_y1 + _y2 + _y3) / 3
        if _avg == 0:
            fields["rd_credit_amount"] = _cq * 0.06
        else:
            fields["rd_credit_amount"] = max(0.0, (_cq - _avg * 0.50) * 0.14)

    # Sheet data
    company_rows = read_qa_sheet(company_overview_ws) if company_overview_ws else []
    project_rows = read_qa_sheet(project_ws) if project_ws else []

    wage_rows, total_taxable, total_qre_w = (
        read_wage_qre(wage_ws) if wage_ws else ([], 0.0, 0.0))
    supply_rows, total_qre_s = (
        read_supply_qre(supply_ws) if supply_ws else ([], 0.0))
    contractor_us_rows, total_qre_us = (
        read_contractor_sheet(contractor_us_ws) if contractor_us_ws else ([], 0.0))
    contractor_for_rows, total_qre_for = (
        read_contractor_sheet(contractor_for_ws) if contractor_for_ws else ([], 0.0))

    tax_year_int = int(str(fields.get("tax_year", "2022")).strip())
    col_headers, receipts_data, receipts_total = (
        read_gross_receipts(receipts_ws, tax_year_int) if receipts_ws
        else (["Income Item", str(tax_year_int-2), str(tax_year_int-1)], [], None))
    comment_rows    = read_comments(comments_ws)   if comments_ws   else []
    credit_tab_rows = read_credit_tab(credit_tab_ws) if credit_tab_ws else []

    # Narrative txt files
    company_paras = read_text_file(input_dir / "company_overview.txt")
    project_paras = read_text_file(input_dir / "project_overview.txt")
    if not company_paras: print("WARNING: Input/company_overview.txt not found.")
    if not project_paras: print("WARNING: Input/project_overview.txt not found.")

    # Build pages
    cover_html          = build_cover_pages(fields)
    credit_tab_html     = build_credit_tab_page(credit_tab_rows, fields)
    company_html        = build_narrative_page("Company Overview", company_paras, fields)
    project_html        = build_narrative_page("Project Overview", project_paras, fields)
    wage_html           = build_wage_page(wage_rows, total_taxable, total_qre_w, fields)
    supply_html         = build_supply_page(supply_rows, total_qre_s, fields)
    contractor_us_html  = build_contractor_page("Contractor US",      contractor_us_rows,  total_qre_us,  "State",   fields)
    contractor_for_html = build_contractor_page("Contractor Foreign", contractor_for_rows, total_qre_for, "Country", fields)
    receipts_html       = build_receipts_page(col_headers, receipts_data, receipts_total, fields)
    comments_html       = build_comments_page(comment_rows, fields)

    pages_html = (cover_html + credit_tab_html + company_html + project_html
                  + wage_html + supply_html
                  + contractor_us_html + contractor_for_html
                  + receipts_html + comments_html)

    company_raw = str(fields.get("company_name","Client")).replace(" ","_")
    tax_yr_raw  = str(fields.get("tax_year","")).strip()
    out_name    = "rd_study_" + company_raw + "_" + tax_yr_raw + ".html"
    out_path    = output_dir / out_name

    title_esc = "R&amp;D Tax Credit Study"
    html_doc = ("<!DOCTYPE html>\n<html lang=\"en\">\n<head>\n"
                "<meta charset=\"UTF-8\">\n"
                "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">\n"
                "<title>" + title_esc + " — " + esc(company_raw) + " " + esc(tax_yr_raw) + "</title>\n"
                "<style>\n" + CSS + "\n</style>\n"
                "</head>\n<body>\n"
                + pages_html
                + "</body>\n</html>")

    out_path.write_text(html_doc, encoding="utf-8")
    print("\nReport saved: " + str(out_path))
    print("Open in your browser, then File -> Print -> Save as PDF")


if __name__ == "__main__":
    main()
